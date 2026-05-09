import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

gastos = [
  # ENERO
  ('01/01/2026','Gasto','Cuenta de Prueba','Educación','Servicios','ACADEMIA JULIETA',-1900),
  ('01/01/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL GABINO 098927294',-4425),
  ('24/01/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL INTERNET',-2022.30),
  ('24/01/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL MATEO 091756286',-1491.15),
  ('24/01/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL PAULINA 091678144',-1543.15),
  ('30/01/2026','Gasto','Cuenta de Prueba','CJPPU','Servicios','CAJA PROFESIONAL',-23316),
  ('01/01/2026','Gasto','Cuenta de Prueba','IM','Servicios','ESTACIONAMIENTO',-276),
  ('20/01/2026','Gasto','Cuenta de Prueba','OSE','Servicios','OSE',-1071.30),
  ('15/01/2026','Gasto','Cuenta de Prueba','ADT','Servicios','SEGURIDAD',-4525.71),
  ('21/01/2026','Gasto','Cuenta de Prueba','UTE','Servicios','UTE',-16231.03),
  # FEBRERO
  ('01/02/2026','Gasto','Cuenta de Prueba','Educación','Servicios','ACADEMIA JULIETA',-1900),
  ('01/02/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL GABINO 098927294',-4425),
  ('24/02/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL INTERNET',-2022.30),
  ('24/02/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL MATEO 091756286',-1491.15),
  ('24/02/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL PAULINA 091678144',-1543.15),
  ('28/02/2026','Gasto','Cuenta de Prueba','CJPPU','Servicios','CAJA PROFESIONAL',-23316),
  ('01/02/2026','Gasto','Cuenta de Prueba','IM','Servicios','ESTACIONAMIENTO',-276),
  ('20/02/2026','Gasto','Cuenta de Prueba','OSE','Servicios','OSE',-1071.30),
  ('15/02/2026','Gasto','Cuenta de Prueba','ADT','Servicios','SEGURIDAD',-4525.71),
  ('21/02/2026','Gasto','Cuenta de Prueba','UTE','Servicios','UTE',-16231.03),
  # MARZO
  ('01/03/2026','Gasto','Cuenta de Prueba','Educación','Servicios','ACADEMIA JULIETA',-1900),
  ('01/03/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL GABINO 098927294',-4121),
  ('24/03/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL INTERNET',-2102.96),
  ('24/03/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL MATEO 091756286',-1640.66),
  ('24/03/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL PAULINA 091678144',-1640.66),
  ('24/03/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL JULIETA',-12137.64),
  ('30/03/2026','Gasto','Cuenta de Prueba','CJPPU','Servicios','CAJA PROFESIONAL',-27081),
  ('01/03/2026','Gasto','Cuenta de Prueba','IM','Servicios','ESTACIONAMIENTO',-278),
  ('20/03/2026','Gasto','Cuenta de Prueba','OSE','Servicios','OSE',-1328.10),
  ('15/03/2026','Gasto','Cuenta de Prueba','ADT','Servicios','SEGURIDAD',-4451.52),
  ('21/03/2026','Gasto','Cuenta de Prueba','UTE','Servicios','UTE',-15434.88),
  # ABRIL
  ('01/04/2026','Gasto','Cuenta de Prueba','Educación','Servicios','ACADEMIA JULIETA',-1900),
  ('01/04/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL GABINO 098927294',-4003),
  ('24/04/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL INTERNET',-2372.96),
  ('24/04/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL MATEO 091756286',-1487.22),
  ('24/04/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL PAULINA 091678144',-1487.22),
  ('24/04/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL JULIETA',-3857.89),
  ('30/04/2026','Gasto','Cuenta de Prueba','CJPPU','Servicios','CAJA PROFESIONAL',-27081),
  ('01/04/2026','Gasto','Cuenta de Prueba','IM','Servicios','ESTACIONAMIENTO',-287),
  ('20/04/2026','Gasto','Cuenta de Prueba','OSE','Servicios','OSE',-1051.24),
  ('21/04/2026','Gasto','Cuenta de Prueba','UTE','Servicios','UTE',-14792.31),
  # MAYO
  ('01/05/2026','Gasto','Cuenta de Prueba','Educación','Servicios','ACADEMIA JULIETA',-1900),
  ('01/05/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL GABINO 098927294',-2102.96),
  ('24/05/2026','Gasto','Cuenta de Prueba','Antel','Servicios','ANTEL INTERNET',-2102.96),
]

# Ingresos inventados: mayores que gastos del mes, diferentes cada mes, mayo=febrero
ingresos = [
  ('05/01/2026','Ingreso','Cuenta de Prueba','Sueldo Base','Sueldo Nominal','Salario enero 2026', 68000),
  ('05/02/2026','Ingreso','Cuenta de Prueba','Sueldo Base','Sueldo Nominal','Salario febrero 2026', 75000),
  ('05/03/2026','Ingreso','Cuenta de Prueba','Sueldo Base','Sueldo Nominal','Salario marzo 2026', 90000),
  ('05/04/2026','Ingreso','Cuenta de Prueba','Sueldo Base','Sueldo Nominal','Salario abril 2026', 72000),
  ('05/05/2026','Ingreso','Cuenta de Prueba','Sueldo Base','Sueldo Nominal','Salario mayo 2026', 75000),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Fluxo Import'

# Filas 1-2 vacias (formato Zcuentas)
ws.append([])
ws.append([])

# Fila 3: headers
headers = ['Fecha', 'Tipo', 'Cuenta', 'Concepto', 'Etiqueta', 'Descripcion', 'Importe']
ws.append(headers)

header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_font = Font(color='10B981', bold=True)
for cell in ws[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for row in ingresos + gastos:
    ws.append(list(row))

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 12

out = r'C:\Users\Luis Enrique\Desktop\GABINO_FLUXO.xlsx'
wb.save(out)
print('Generado:', out)
print('Total filas:', len(ingresos) + len(gastos))
print('Ingresos:', len(ingresos), '| Gastos:', len(gastos))
print()
print('Resumen ingresos vs gastos por mes:')
meses = {1:0,2:0,3:0,4:0,5:0}
for g in gastos:
    mes = int(g[0].split('/')[1])
    meses[mes] += abs(g[6])
ing = {1:68000,2:75000,3:90000,4:72000,5:75000}
for m in range(1,6):
    print(f'  Mes {m}: Ingresos={ing[m]:,.0f} | Gastos={meses[m]:,.2f} | Ahorro={ing[m]-meses[m]:,.2f}')
